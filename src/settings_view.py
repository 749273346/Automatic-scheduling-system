from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, 
                             QPushButton, QHeaderView, QInputDialog, QColorDialog, QMessageBox, QLabel, QSpinBox, QGroupBox,
                             QDialog, QTabWidget, QCalendarWidget, QCheckBox, QComboBox, QLineEdit, QFormLayout, QDialogButtonBox, QSpacerItem, QSizePolicy,
                             QScrollArea, QGridLayout, QListWidget, QListWidgetItem, QMenu, QAction, QFileDialog, QProgressDialog)
from PyQt5.QtCore import Qt, QLocale
import openpyxl
import random
from src.models import User
from src.db_manager import DBManager

class SettingsView(QWidget):
    def __init__(self, users, db_manager: DBManager, main_window):
        super().__init__()
        self.users = users
        self.db_manager = db_manager
        self.main_window = main_window # Reference to main window to update UI if needed
        self.preferences_msg_shown = False
        self.init_ui()

    def reset_msg_state(self):
        self.preferences_msg_shown = False

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)

        # Title
        title = QLabel("äººå‘˜ç®¡ç†è®¾ç½®")
        title.setObjectName("SettingsTitle")
        layout.addWidget(title)
        
        # --- Action Bar ---
        action_layout = QHBoxLayout()
        
        # Search
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("ğŸ” æœç´¢äººå‘˜...")
        self.search_input.setFixedWidth(250)
        self.search_input.setStyleSheet("padding: 5px; border-radius: 15px; border: 1px solid #ddd;")
        self.search_input.textChanged.connect(self.load_users)
        action_layout.addWidget(self.search_input)

        action_layout.addStretch()

        # Clear Preferences Button
        self.btn_clear_prefs = QPushButton("æ¸…é™¤æ‰€æœ‰åå¥½")
        self.btn_clear_prefs.setCursor(Qt.PointingHandCursor)
        self.btn_clear_prefs.clicked.connect(self.clear_all_preferences)
        self.btn_clear_prefs.setStyleSheet("""
            QPushButton {
                background-color: white;
                color: #FF3B30;
                border: 1px solid #FF3B30;
                border-radius: 6px;
                padding: 6px 15px;
                font-weight: bold;
                margin-right: 10px;
            }
            QPushButton:hover {
                background-color: #FFF0F0;
            }
        """)
        action_layout.addWidget(self.btn_clear_prefs)

        # Excel Import Button (New)
        self.btn_import = QPushButton(" Excelè‡ªåŠ¨å¯¼å…¥ ")
        self.btn_import.setCursor(Qt.PointingHandCursor)
        self.btn_import.clicked.connect(self.import_from_excel)
        self.btn_import.setStyleSheet("""
            QPushButton {
                background-color: #34C759;
                color: white;
                border-radius: 6px;
                padding: 6px 15px;
                font-weight: bold;
                margin-right: 10px;
            }
            QPushButton:hover {
                background-color: #2da84e;
            }
        """)
        action_layout.addWidget(self.btn_import)

        # Add User Button
        self.btn_add = QPushButton(" + æ·»åŠ äººå‘˜ ")
        self.btn_add.setStyleSheet("""
            QPushButton {
                background-color: #007AFF;
                color: white;
                border-radius: 6px;
                padding: 6px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0069D9;
            }
        """)
        self.btn_add.setCursor(Qt.PointingHandCursor)
        self.btn_add.clicked.connect(self.add_user)
        action_layout.addWidget(self.btn_add)
        
        layout.addLayout(action_layout)
        
        # --- Staff List Section ---
        layout.addSpacing(10)

        # Table
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget::item:selected {
                background-color: #E5F3FF;
                color: black;
            }
        """)
        self.table.setShowGrid(False)
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["ID (Code)", "å§“å", "èŒä½", "è”ç³»æ–¹å¼", "é¢œè‰²", "æ“ä½œ"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSortingEnabled(True) # Enable sorting
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        layout.addWidget(self.table)
        
        # Empty State Label
        self.lbl_empty = QLabel("æš‚æ— äººå‘˜æ•°æ®ï¼Œè¯·ç‚¹å‡»å³ä¸Šè§’æ·»åŠ æŒ‰é’®")
        self.lbl_empty.setAlignment(Qt.AlignCenter)
        self.lbl_empty.setStyleSheet("color: #888; font-size: 16px; margin: 20px;")
        self.lbl_empty.setVisible(False)
        layout.addWidget(self.lbl_empty)

        # Footer Actions
        footer_layout = QHBoxLayout()
        
        self.lbl_count = QLabel()
        footer_layout.addWidget(self.lbl_count)
        
        footer_layout.addStretch()
        
        # System Reset (Kept small)
        self.btn_reset = QPushButton("ç³»ç»Ÿé‡ç½®")
        self.btn_reset.setFlat(True)
        self.btn_reset.setStyleSheet("color: #999; text-decoration: underline;")
        self.btn_reset.setCursor(Qt.PointingHandCursor)
        self.btn_reset.clicked.connect(self.reset_system)
        footer_layout.addWidget(self.btn_reset)
        
        self.btn_refresh = QPushButton("åˆ·æ–°")
        self.btn_refresh.clicked.connect(self.load_users)
        footer_layout.addWidget(self.btn_refresh)
        
        layout.addLayout(footer_layout)
        
        # Initial spin count for reset logic compatibility
        self.spin_count = QSpinBox()
        self.spin_count.setValue(8) 
        self.spin_count.setVisible(False) # Hidden but kept for reset_system logic if needed

        self.load_users()

    def show_context_menu(self, pos):
        """Show context menu on right click"""
        # Get selected rows
        selected_rows = sorted(set(index.row() for index in self.table.selectionModel().selectedRows()))
        
        item = self.table.itemAt(pos)
        
        # If right-click happens on an item not in current selection, treat it as single item action
        # (unless user Ctrl+Click, but right click usually implies context of "what is under cursor" or "current selection")
        # Standard behavior: If click is inside selection, apply to selection. If outside, apply to that item (and usually select it).
        
        clicked_on_selection = False
        if item:
            if item.row() in selected_rows:
                clicked_on_selection = True
        
        menu = QMenu(self)
        
        # Batch Operation if multiple rows selected AND clicked on selection
        if len(selected_rows) > 1 and clicked_on_selection:
            delete_action = QAction(f"æ‰¹é‡åˆ é™¤ ({len(selected_rows)} äºº)", self)
            delete_action.triggered.connect(lambda: self.delete_selected_users(selected_rows))
            menu.addAction(delete_action)
            
        else:
            # Single item operation
            if not item:
                return
                
            row = item.row()
            user_item = self.table.item(row, 0)
            if not user_item:
                return
                
            user = user_item.data(Qt.UserRole)
            if not user:
                return
            
            delete_action = QAction("åˆ é™¤äººå‘˜", self)
            delete_action.triggered.connect(lambda: self.delete_user(user))
            menu.addAction(delete_action)
            
            edit_action = QAction("ç¼–è¾‘äººå‘˜", self)
            edit_action.triggered.connect(lambda: self.edit_user(user))
            menu.addAction(edit_action)
            
            pref_action = QAction("åå¥½è®¾ç½®", self)
            pref_action.triggered.connect(lambda: self.edit_preferences(user))
            menu.addAction(pref_action)
            
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def delete_selected_users(self, rows):
        """Delete multiple users"""
        users_to_delete = []
        for row in rows:
            item = self.table.item(row, 0)
            if item:
                user = item.data(Qt.UserRole)
                if user:
                    users_to_delete.append(user)
        
        if not users_to_delete:
            return

        names = ", ".join([u.code for u in users_to_delete[:5]])
        if len(users_to_delete) > 5:
            names += " ç­‰"
            
        reply = QMessageBox.question(self, "ç¡®è®¤æ‰¹é‡åˆ é™¤", 
                                     f"ç¡®å®šè¦åˆ é™¤ä»¥ä¸‹ {len(users_to_delete)} ä½äººå‘˜å—ï¼Ÿ\n{names}\n\næ­¤æ“ä½œä¸å¯æ¢å¤ã€‚",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            success_count = 0
            # Use DB transaction if possible, or just loop
            # Here we loop for simplicity as existing DBManager handles single deletes safely
            for user in users_to_delete:
                if self.db_manager.delete_user(user.id):
                    success_count += 1
            
            if success_count > 0:
                if hasattr(self.main_window, 'reload_data'):
                    self.main_window.reload_data()
                else:
                    self.load_users() # Fallback
                QMessageBox.information(self, "æˆåŠŸ", f"æˆåŠŸåˆ é™¤ {success_count} ä½äººå‘˜")
            else:
                QMessageBox.warning(self, "å¤±è´¥", "åˆ é™¤å¤±è´¥")

    def load_users(self):
        # Disable sorting while loading to prevent artifacts
        self.table.setSortingEnabled(False)
        
        search_text = self.search_input.text().strip().lower()
        filtered_users = []
        for u in self.users:
            # Filter logic
            if search_text:
                if (search_text not in u.code.lower() and 
                    search_text not in (u.name or "").lower() and 
                    search_text not in (u.position or "").lower()):
                    continue
            filtered_users.append(u)
            
        self.table.setRowCount(0)
        self.lbl_count.setText(f"æ˜¾ç¤º {len(filtered_users)} / {len(self.users)} äºº")
        
        if not filtered_users:
            self.table.setVisible(False)
            self.lbl_empty.setVisible(True)
            if search_text:
                self.lbl_empty.setText("æœªæ‰¾åˆ°åŒ¹é…çš„äººå‘˜")
            else:
                self.lbl_empty.setText("æš‚æ— äººå‘˜æ•°æ®ï¼Œè¯·ç‚¹å‡»å³ä¸Šè§’æ·»åŠ æŒ‰é’®")
        else:
            self.table.setVisible(True)
            self.lbl_empty.setVisible(False)
        
        for row, user in enumerate(filtered_users):
            self.table.insertRow(row)
            
            # ID
            item_code = QTableWidgetItem(user.code)
            item_code.setTextAlignment(Qt.AlignCenter)
            item_code.setData(Qt.UserRole, user) # Store user object if needed
            self.table.setItem(row, 0, item_code)
            
            # Name
            item_name = QTableWidgetItem(user.name or "")
            item_name.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 1, item_name)
            
            # Position
            item_pos = QTableWidgetItem(user.position or "-")
            item_pos.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 2, item_pos)
            
            # Contact
            item_contact = QTableWidgetItem(user.contact or "-")
            item_contact.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 3, item_contact)

            # Color
            btn_color = QPushButton()
            btn_color.setStyleSheet(f"background-color: {user.color}; border: 1px solid #ddd; border-radius: 4px;")
            btn_color.setCursor(Qt.PointingHandCursor)
            btn_color.clicked.connect(lambda checked, u=user: self.change_color(u))
            btn_color.setFixedSize(60, 20)
            
            # Center widget in cell
            color_container = QWidget()
            color_layout = QHBoxLayout(color_container)
            color_layout.setContentsMargins(0, 0, 0, 0)
            color_layout.setAlignment(Qt.AlignCenter)
            color_layout.addWidget(btn_color)
            self.table.setCellWidget(row, 4, color_container)

            # Actions
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(0, 2, 0, 2)
            action_layout.setAlignment(Qt.AlignCenter)
            
            btn_edit = QPushButton("ç¼–è¾‘")
            btn_edit.setCursor(Qt.PointingHandCursor)
            btn_edit.setStyleSheet("color: #007AFF; border: none; font-weight: bold;")
            btn_edit.clicked.connect(lambda checked, u=user: self.edit_user(u))
            action_layout.addWidget(btn_edit)
            
            btn_pref = QPushButton("åå¥½")
            btn_pref.setCursor(Qt.PointingHandCursor)
            btn_pref.setStyleSheet("color: #5856D6; border: none;")
            btn_pref.clicked.connect(lambda checked, u=user: self.edit_preferences(u))
            action_layout.addWidget(btn_pref)
            
            btn_del = QPushButton("åˆ é™¤")
            btn_del.setCursor(Qt.PointingHandCursor)
            btn_del.setStyleSheet("color: #FF3B30; border: none;")
            btn_del.clicked.connect(lambda checked, u=user: self.delete_user(u))
            action_layout.addWidget(btn_del)
            
            self.table.setCellWidget(row, 5, action_widget)
            
        self.table.setSortingEnabled(True) # Re-enable sorting
    
    def clear_all_preferences(self):
        reply = QMessageBox.question(
            self, 
            "ç¡®è®¤æ¸…é™¤", 
            "ç¡®å®šè¦æ¸…é™¤æ‰€æœ‰äººå‘˜çš„åå¥½è®¾ç½®å—ï¼Ÿ\næ­¤æ“ä½œå°†é‡ç½®æ‰€æœ‰äººçš„æ’ç­åå¥½ï¼ˆå¦‚ä¸å¯æ’ç­æ—¥æœŸã€åå¥½å·¥ä½œæ—¥ç­‰ï¼‰ã€‚\n\næ³¨æ„ï¼šæ­¤æ“ä½œä¸å¯æ’¤é”€ã€‚",
            QMessageBox.Yes | QMessageBox.No, 
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self.db_manager.clear_all_preferences():
                if hasattr(self.main_window, 'reload_data'):
                    self.main_window.reload_data()
                else:
                    self.users = self.db_manager.get_all_users()
                    self.load_users()
                QMessageBox.information(self, "æˆåŠŸ", "æ‰€æœ‰åå¥½è®¾ç½®å·²æ¸…é™¤ã€‚")
            else:
                QMessageBox.warning(self, "å¤±è´¥", "æ¸…é™¤åå¥½è®¾ç½®å¤±è´¥ï¼Œè¯·é‡è¯•ã€‚")

    def import_from_excel(self):
        """Import users from Excel file with smart column recognition"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "é€‰æ‹©äººå‘˜ä¿¡æ¯è¡¨", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
            
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            # 1. Identify Header Row
            header_row_idx = None
            column_map = {}
            
            # Keywords for column mapping
            keywords = {
                'code': ['id', 'code', 'ç¼–å·', 'å·¥å·', 'ä»£ç '],
                'name': ['å§“å', 'åå­—', 'name'],
                'position': ['èŒåŠ¡', 'èŒä½', 'å²—ä½', 'role', 'position'],
                'contact': ['è”ç³»æ–¹å¼', 'ç”µè¯', 'æ‰‹æœº', 'contact', 'phone', 'tel'],
                'color': ['é¢œè‰²', 'color'],
                'priority': ['ä¼˜å…ˆç­‰çº§', 'ç­‰çº§', 'å‘˜å·¥ç±»å‹', 'priority', 'level', 'type']
            }
            
            # Scan first 10 rows for headers
            for r in range(1, min(11, sheet.max_row + 1)):
                row_values = [str(sheet.cell(row=r, column=c).value or "").strip().lower() for c in range(1, sheet.max_column + 1)]
                
                # Check if this row looks like a header (contains at least 'name' or 'code')
                matches = 0
                temp_map = {}
                
                for col_idx, cell_val in enumerate(row_values):
                    # Check against keywords
                    for key, words in keywords.items():
                        if any(w in cell_val for w in words):
                            temp_map[key] = col_idx + 1 # 1-based column index
                            break
                            
                if 'name' in temp_map or 'code' in temp_map:
                    if len(temp_map) >= 2: # At least 2 columns matched
                        header_row_idx = r
                        column_map = temp_map
                        break
            
            if not header_row_idx:
                QMessageBox.warning(self, "è¯†åˆ«å¤±è´¥", "æ— æ³•è¯†åˆ«è¡¨å¤´ï¼Œè¯·ç¡®ä¿è¡¨æ ¼åŒ…å«'å§“å'ã€'å·¥å·'ç­‰åˆ—åã€‚")
                return
                
            # 2. Process Data Rows
            success_count = 0
            fail_count = 0
            errors = []
            
            # Pre-fetch existing codes to avoid duplicates
            existing_codes = {u.code.strip().upper() for u in self.db_manager.get_all_users()}
            
            # Pre-scan Excel for explicit codes to ensure generator doesn't conflict
            excel_explicit_codes = set()
            for r in range(header_row_idx + 1, sheet.max_row + 1):
                if 'code' in column_map:
                    val = sheet.cell(row=r, column=column_map['code']).value
                    if val:
                        code_str = str(val).strip().upper()
                        if code_str:
                            excel_explicit_codes.add(code_str)
                            
            used_codes = existing_codes.union(excel_explicit_codes)
            
            # Helper to generate random color
            def generate_random_color():
                """Generate a random pleasing color"""
                # Generate RGB values ensuring they aren't too dark or too light
                r = random.randint(60, 220)
                g = random.randint(60, 220)
                b = random.randint(60, 220)
                return f"#{r:02X}{g:02X}{b:02X}"

            # Helper to generate next available code
            def generate_next_code():
                # Try single letters A-Z
                for i in range(26):
                    c = chr(65 + i)
                    if c not in used_codes:
                        used_codes.add(c)
                        return c
                # Try double letters AA-ZZ
                for i in range(26):
                    for j in range(26):
                        c = chr(65 + i) + chr(65 + j)
                        if c not in used_codes:
                            used_codes.add(c)
                            return c
                # Fallback numeric
                idx = 1
                while True:
                    c = f"U{idx}"
                    if c not in used_codes:
                        used_codes.add(c)
                        return c
                    idx += 1
            
            # Progress Dialog
            progress = QProgressDialog("æ­£åœ¨å¯¼å…¥æ•°æ®...", "å–æ¶ˆ", 0, sheet.max_row - header_row_idx, self)
            progress.setWindowModality(Qt.WindowModal)
            
            for i, r in enumerate(range(header_row_idx + 1, sheet.max_row + 1)):
                if progress.wasCanceled():
                    break
                
                progress.setValue(i)
                
                # Extract values
                def get_val(key):
                    if key in column_map:
                        val = sheet.cell(row=r, column=column_map[key]).value
                        return str(val).strip() if val is not None else None
                    return None
                
                code = get_val('code')
                name = get_val('name')
                
                # Skip empty rows (neither name nor code)
                if not code and not name:
                    continue
                    
                # Auto-generate ID if missing
                if not code:
                     code = generate_next_code()
                     # If name is also missing (should be caught above), but double check
                     if not name:
                         name = f"å‘˜å·¥{code}"
                
                position = get_val('position')
                contact = get_val('contact')
                color = get_val('color')
                if not color:
                    color = generate_random_color()
                priority_val = get_val('priority')
                
                # Parse priority
                prefs = {}
                if priority_val:
                    if "ä¸€" in priority_val or "1" in priority_val:
                        prefs["employee_type"] = "ä¸€çº§"
                    elif "äºŒ" in priority_val or "2" in priority_val:
                        prefs["employee_type"] = "äºŒçº§"
                    elif "ä¸‰" in priority_val or "3" in priority_val:
                        prefs["employee_type"] = "ä¸‰çº§"
                    else:
                         prefs["employee_type"] = "ä¸€çº§" # Default
                
                # Add to DB
                user, msg = self.db_manager.add_user(
                    code=code,
                    name=name,
                    position=position,
                    contact=contact,
                    color=color,
                    preferences=prefs
                )
                
                if user:
                    success_count += 1
                else:
                    # If it failed because ID exists, maybe update?
                    # "å‘˜å·¥ä»£ç (ID)å·²å­˜åœ¨"
                    if "å­˜åœ¨" in msg:
                        # Update logic could go here if requested
                        # For now, just report error
                        fail_count += 1
                        errors.append(f"è¡Œ {r} ({name}): {msg}")
                    else:
                        fail_count += 1
                        errors.append(f"è¡Œ {r} ({name}): {msg}")
            
            progress.setValue(sheet.max_row - header_row_idx)
            
            # Reload UI
            self.users = self.db_manager.get_all_users()
            self.load_users()
            if hasattr(self.main_window, 'reload_data'):
                self.main_window.reload_data()
                
            # Report
            if fail_count == 0:
                QMessageBox.information(self, "å¯¼å…¥å®Œæˆ", f"æˆåŠŸå¯¼å…¥ {success_count} æ¡æ•°æ®ã€‚")
            else:
                err_msg = "\n".join(errors[:10])
                if len(errors) > 10:
                    err_msg += "\n..."
                QMessageBox.warning(self, "å¯¼å…¥å®Œæˆ (æœ‰é”™è¯¯)", f"æˆåŠŸ: {success_count}\nå¤±è´¥: {fail_count}\n\né”™è¯¯è¯¦æƒ…:\n{err_msg}")
                
        except Exception as e:
            QMessageBox.critical(self, "å¯¼å…¥é”™è¯¯", f"è¯»å–æ–‡ä»¶æ—¶å‘ç”Ÿé”™è¯¯: {str(e)}")

    def add_user(self):
        dialog = UserEditDialog(parent=self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            
            # Prepare preferences
            prefs = {}
            if "employee_type" in data:
                prefs["employee_type"] = data["employee_type"]

            user, msg = self.db_manager.add_user(
                code=data['code'],
                name=data['name'],
                position=data['position'],
                contact=data['contact'],
                color=data['color'],
                preferences=prefs
            )
            if user:
                # Refresh data
                self.users.append(user)
                self.load_users()
                if hasattr(self.main_window, 'reload_data'):
                    self.main_window.reload_data()
                
                if not self.preferences_msg_shown:
                    QMessageBox.information(self, "æˆåŠŸ", "äººå‘˜æ·»åŠ æˆåŠŸ")
                    self.preferences_msg_shown = True
            else:
                QMessageBox.warning(self, "å¤±è´¥", f"æ·»åŠ å¤±è´¥: {msg}")

    def edit_user(self, user):
        dialog = UserEditDialog(user, parent=self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            
            # Prepare preferences update
            current_prefs = dict(user.preferences) if user.preferences else {}
            current_prefs["employee_type"] = data.get("employee_type", "ä¸€çº§")
            
            # Code is usually immutable or needs check
            # Update DB
            success, msg = self.db_manager.update_user(
                user.id,
                name=data['name'],
                position=data['position'],
                contact=data['contact'],
                color=data['color'],
                preferences=current_prefs
            )
            if success:
                # Update memory
                user.name = data['name']
                user.position = data['position']
                user.contact = data['contact']
                user.color = data['color']
                user.preferences = current_prefs
                self.load_users()
                if hasattr(self.main_window, 'reload_data'):
                    self.main_window.reload_data()
                
                if not self.preferences_msg_shown:
                    QMessageBox.information(self, "æˆåŠŸ", "äººå‘˜ä¿¡æ¯å·²æ›´æ–°\n\næ³¨æ„ï¼šä¿®æ”¹ä»…å½±å“åç»­è‡ªåŠ¨æ’ç­ï¼Œç°æœ‰æ’ç­è®°å½•ä¸ä¼šæ”¹å˜ã€‚")
                    self.preferences_msg_shown = True
            else:
                QMessageBox.warning(self, "å¤±è´¥", f"æ›´æ–°å¤±è´¥: {msg}")

    def delete_user(self, user):
        reply = QMessageBox.question(self, "ç¡®è®¤åˆ é™¤", 
                                     f"ç¡®å®šè¦åˆ é™¤äººå‘˜ {user.name or user.code} å—ï¼Ÿ\næ­¤æ“ä½œä¸å¯æ¢å¤ã€‚\n\næ³¨æ„ï¼šåˆ é™¤åï¼Œè¯¥äººå‘˜çš„å†å²æ’ç­è®°å½•å°†ä¿ç•™ï¼Œä½†ä¸ä¼šå†å‚ä¸æ–°çš„æ’ç­ã€‚",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            if self.db_manager.delete_user(user.id):
                self.users = [u for u in self.users if u.id != user.id]
                self.load_users()
                if hasattr(self.main_window, 'reload_data'):
                    self.main_window.reload_data()
                
                if not self.preferences_msg_shown:
                    QMessageBox.information(self, "æˆåŠŸ", "äººå‘˜å·²åˆ é™¤")
                    self.preferences_msg_shown = True
            else:
                QMessageBox.warning(self, "å¤±è´¥", "åˆ é™¤å¤±è´¥")


    def change_color(self, user):
        color = QColorDialog.getColor(initial=Qt.white, parent=self, title=f"é€‰æ‹© {user.code} çš„é¢œè‰²")
        if color.isValid():
            new_color = color.name()
            # Update DB
            session = self.db_manager.get_session()
            db_user = session.query(User).filter_by(id=user.id).first()
            if db_user:
                db_user.color = new_color
                session.commit()
                # Update memory object
                user.color = new_color
                # Refresh UI
                self.load_users()
                if hasattr(self.main_window, 'reload_data'):
                    self.main_window.reload_data()
                
                if not self.preferences_msg_shown:
                    QMessageBox.information(self, "æˆåŠŸ", "é¢œè‰²å·²æ›´æ–°")
                    self.preferences_msg_shown = True
            session.close()

    def edit_preferences(self, user):
        dialog = PreferenceDialog(user, self.users, self)
        if dialog.exec_() == QDialog.Accepted:
            new_prefs = dialog.get_preferences()
            
            # Update DB
            session = self.db_manager.get_session()
            db_user = session.query(User).filter_by(id=user.id).first()
            if db_user:
                db_user.preferences = new_prefs
                # Force SQLAlchemy to detect change in JSON column
                from sqlalchemy.orm.attributes import flag_modified
                flag_modified(db_user, "preferences")
                
                session.commit()
                
                # Update memory object
                user.preferences = new_prefs
                pr = new_prefs.get("periodic_rotation")
                if pr and pr.get("partner") and pr.get("day_idx") is not None:
                    partner_code = pr["partner"]
                    partner = next((u for u in self.users if u.code == partner_code), None)
                    if partner:
                        pdb = session.query(User).filter_by(id=partner.id).first()
                        if pdb:
                            pp = dict(pdb.preferences) if pdb.preferences else {}
                            pw = set(pp.get("preferred_weekdays", []))
                            pw.add(pr["day_idx"])
                            pp["preferred_weekdays"] = sorted(list(pw))
                            pp["periodic_rotation"] = {
                                "partner": user.code,
                                "day_idx": pr["day_idx"],
                                "parity": "even" if pr.get("parity", "odd") == "odd" else "odd"
                            }
                            pdb.preferences = pp
                            flag_modified(pdb, "preferences")
                            session.commit()
                            partner.preferences = pp
                if not self.preferences_msg_shown:
                    QMessageBox.information(self, "æˆåŠŸ", "åå¥½è®¾ç½®å·²ä¿å­˜\n\næ³¨æ„ï¼šæ–°çš„åå¥½è®¾ç½®å°†ä»…åº”ç”¨äºåç»­ç”Ÿæˆçš„æ’ç­ï¼Œç°æœ‰æ’ç­ä¸ä¼šå—å½±å“ã€‚")
                    self.preferences_msg_shown = True
            session.close()

    def reset_system(self):
        count = self.spin_count.value()
        reply = QMessageBox.question(self, "ç¡®è®¤é‡ç½®", 
                                     f"ç¡®å®šè¦é‡ç½®ç³»ç»Ÿä¸º {count} äººå—ï¼Ÿ\nè­¦å‘Šï¼šè¿™å°†æ¸…ç©ºæ‰€æœ‰ç°æœ‰çš„æ’ç­è®°å½•ï¼",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                self.db_manager.reset_users(count)
                # Refresh Main Window Data
                self.main_window.reload_data()
                QMessageBox.information(self, "é‡ç½®æˆåŠŸ", f"ç³»ç»Ÿå·²é‡ç½®ä¸º {count} äººã€‚")
            except Exception as e:
                QMessageBox.critical(self, "é”™è¯¯", f"é‡ç½®å¤±è´¥: {str(e)}")

    def update_data(self, users):
        self.users = users
        self.spin_count.setValue(len(users))
        self.load_users()


class PreferenceDialog(QDialog):
    def __init__(self, user, all_users, parent=None):
        super().__init__(parent)
        self.user = user
        self.all_users = all_users
        self.preferences = dict(user.preferences) if user.preferences else {}
        self.lock_rotation = None
        for u in self.all_users:
            if not u.preferences:
                continue
            pr = u.preferences.get("periodic_rotation")
            if pr and pr.get("partner") == self.user.code:
                self.lock_rotation = {
                    "partner": u.code,
                    "day_idx": pr.get("day_idx", 4),
                    "parity": "even" if pr.get("parity", "odd") == "odd" else "odd"
                }
                break
        
        self.setWindowTitle(f"åå¥½è®¾ç½® - {user.code}")
        self.resize(600, 500)
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Tabs
        tabs = QTabWidget()
        layout.addWidget(tabs)
        
        # Tab 1: Preferences (Cycle & Pairing)
        self.tab_advanced = QWidget()
        self.init_advanced_tab()
        tabs.addTab(self.tab_advanced, "æœŸæœ›å€¼ç­æ—¥æœŸ")

        # Tab 3: Blackout Dates
        self.tab_dates = QWidget()
        self.init_dates_tab()
        tabs.addTab(self.tab_dates, "ä¸å¯å€¼ç­æ—¥æœŸ")
        
        # Tab 4: Preferred Partners (New, only effective for Level 1)
        self.tab_partners = QWidget()
        self.init_partners_tab()
        tabs.addTab(self.tab_partners, "æœŸæœ›æ’ç­æ­æ¡£")
        
        # Default to Advanced Preferences (Index 0)
        tabs.setCurrentIndex(0)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_save = QPushButton("ä¿å­˜")
        btn_save.clicked.connect(self.accept)
        btn_cancel = QPushButton("å–æ¶ˆ")
        btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)

    def init_partners_tab(self):
        layout = QVBoxLayout(self.tab_partners)
        
        # Check if current user is Level 1
        current_type = self._get_user_type(self.user)
        if current_type != "ä¸€çº§":
            lbl = QLabel("è¯¥åŠŸèƒ½ä»…é€‚ç”¨äºä¸€çº§äººå‘˜ã€‚")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("color: gray; font-size: 14px; margin-top: 20px;")
            layout.addWidget(lbl)
            layout.addStretch()
            return

        lbl = QLabel("é€‰æ‹©æ‚¨çš„æœŸæœ›æ’ç­æ­æ¡£ (ä»…é™äºŒçº§å’Œä¸‰çº§äººå‘˜):")
        layout.addWidget(lbl)
        
        lbl_hint = QLabel("æç¤ºï¼šå·²é€‰æ­æ¡£å°†æŒ‰åˆ—è¡¨é¡ºåºä¼˜å…ˆå®‰æ’ã€‚")
        lbl_hint.setStyleSheet("color: gray; font-size: 11px;")
        layout.addWidget(lbl_hint)
        
        # Main area: Two lists (Available, Selected)
        h_layout = QHBoxLayout()
        
        # Left: Available
        v_left = QVBoxLayout()
        v_left.addWidget(QLabel("å¾…é€‰äººå‘˜:"))
        self.list_available_partners = QListWidget()
        self.list_available_partners.setSelectionMode(QListWidget.ExtendedSelection)
        v_left.addWidget(self.list_available_partners)
        h_layout.addLayout(v_left)
        
        # Middle: Buttons
        v_mid = QVBoxLayout()
        v_mid.addStretch()
        self.btn_add_partner = QPushButton(">>")
        self.btn_add_partner.clicked.connect(self.add_partner)
        v_mid.addWidget(self.btn_add_partner)
        
        self.btn_remove_partner = QPushButton("<<")
        self.btn_remove_partner.clicked.connect(self.remove_partner)
        v_mid.addWidget(self.btn_remove_partner)
        v_mid.addStretch()
        h_layout.addLayout(v_mid)
        
        # Right: Selected
        v_right = QVBoxLayout()
        v_right.addWidget(QLabel("å·²é€‰æ­æ¡£ (æŒ‰ä¼˜å…ˆçº§æ’åº):"))
        self.list_selected_partners = QListWidget()
        self.list_selected_partners.setSelectionMode(QListWidget.SingleSelection)
        # Enable basic movement, but specific reordering buttons are safer
        v_right.addWidget(self.list_selected_partners)
        
        # Sort buttons for Selected
        h_sort = QHBoxLayout()
        self.btn_move_up = QPushButton("ä¸Šç§»")
        self.btn_move_up.clicked.connect(self.move_partner_up)
        h_sort.addWidget(self.btn_move_up)
        
        self.btn_move_down = QPushButton("ä¸‹ç§»")
        self.btn_move_down.clicked.connect(self.move_partner_down)
        h_sort.addWidget(self.btn_move_down)
        
        v_right.addLayout(h_sort)
        h_layout.addLayout(v_right)
        
        layout.addLayout(h_layout)
        
        # Load Data
        self.load_partners_data()

    def _get_user_type(self, user):
        if not user or not user.preferences:
            return "ä¸€çº§" # Default
        return user.preferences.get("employee_type", "ä¸€çº§")

    def load_partners_data(self):
        self.list_available_partners.clear()
        self.list_selected_partners.clear()
        
        # Get all Level 2 and Level 3 users
        candidates = []
        for u in self.all_users:
            if u.id == self.user.id:
                continue
            u_type = self._get_user_type(u)
            if u_type in ["äºŒçº§", "ä¸‰çº§"]:
                candidates.append(u)
        
        # Existing preferences
        selected_codes = self.preferences.get("preferred_partners", [])
        
        # Populate Selected
        selected_set = set()
        for code in selected_codes:
            # Find user object
            u = next((x for x in candidates if x.code == code), None)
            if u:
                item = QListWidgetItem(f"{u.code} ({u.name or ''})")
                item.setData(Qt.UserRole, u.code)
                self.list_selected_partners.addItem(item)
                selected_set.add(u.code)
        
        # Populate Available
        for u in candidates:
            if u.code not in selected_set:
                item = QListWidgetItem(f"{u.code} ({u.name or ''})")
                item.setData(Qt.UserRole, u.code)
                self.list_available_partners.addItem(item)

    def add_partner(self):
        items = self.list_available_partners.selectedItems()
        if not items:
            return
        for item in items:
            row = self.list_available_partners.row(item)
            self.list_available_partners.takeItem(row)
            self.list_selected_partners.addItem(item)
            
    def remove_partner(self):
        items = self.list_selected_partners.selectedItems()
        if not items:
            return
        for item in items:
            row = self.list_selected_partners.row(item)
            self.list_selected_partners.takeItem(row)
            self.list_available_partners.addItem(item)

    def move_partner_up(self):
        row = self.list_selected_partners.currentRow()
        if row > 0:
            item = self.list_selected_partners.takeItem(row)
            self.list_selected_partners.insertItem(row - 1, item)
            self.list_selected_partners.setCurrentRow(row - 1)

    def move_partner_down(self):
        row = self.list_selected_partners.currentRow()
        if row < self.list_selected_partners.count() - 1:
            item = self.list_selected_partners.takeItem(row)
            self.list_selected_partners.insertItem(row + 1, item)
            self.list_selected_partners.setCurrentRow(row + 1)
        
    def init_dates_tab(self):
        layout = QVBoxLayout(self.tab_dates)
        
        # --- Unavailable Weekdays ---
        lbl_weekdays = QLabel("é€‰æ‹©è¯¥äººå‘˜æ— æ³•å€¼ç­çš„æ˜ŸæœŸ (å›ºå®šæ¯å‘¨ç”Ÿæ•ˆ):")
        layout.addWidget(lbl_weekdays)
        
        grp_weekdays = QGroupBox()
        weekdays_layout = QHBoxLayout(grp_weekdays)
        weekdays_layout.setContentsMargins(10, 5, 10, 5)
        
        self.unavailable_weekday_checks = []
        days = ["å‘¨ä¸€", "å‘¨äºŒ", "å‘¨ä¸‰", "å‘¨å››", "å‘¨äº”", "å‘¨å…­", "å‘¨æ—¥"]
        unavailable_days = self.preferences.get("unavailable_weekdays", [])
        
        for i, day in enumerate(days):
            chk = QCheckBox(day)
            if i in unavailable_days:
                chk.setChecked(True)
            self.unavailable_weekday_checks.append(chk)
            weekdays_layout.addWidget(chk)
            
        layout.addWidget(grp_weekdays)
        layout.addSpacing(10)

        # --- Unavailable Dates ---
        lbl = QLabel("é€‰æ‹©è¯¥äººå‘˜æ— æ³•å€¼ç­çš„æ—¥æœŸï¼ˆç‚¹å‡»æ—¥æœŸåˆ‡æ¢é€‰ä¸­çŠ¶æ€ï¼‰:")
        layout.addWidget(lbl)
        
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.clicked.connect(self.toggle_date)
        layout.addWidget(self.calendar)
        
        lbl_hint = QLabel("å·²é€‰æ—¥æœŸ (å³é”®ç‚¹å‡»åˆ—è¡¨é¡¹å¯åˆ é™¤):")
        layout.addWidget(lbl_hint)

        self.list_dates = QListWidget()
        self.list_dates.setFixedHeight(100)
        self.list_dates.setContextMenuPolicy(Qt.CustomContextMenu)
        self.list_dates.customContextMenuRequested.connect(self.show_date_context_menu)
        layout.addWidget(self.list_dates)
        
        # Load existing blackout dates
        self.blackout_dates = set(self.preferences.get("blackout_dates", []))
        self.update_date_list()
        
    def toggle_date(self, date):
        date_str = date.toString("yyyy-MM-dd")
        if date_str in self.blackout_dates:
            self.blackout_dates.remove(date_str)
        else:
            self.blackout_dates.add(date_str)
        self.update_date_list()
        
    def update_date_list(self):
        self.list_dates.clear()
        if not self.blackout_dates:
            self.list_dates.addItem("æ— ")
            self.list_dates.setEnabled(False)
        else:
            self.list_dates.setEnabled(True)
            sorted_dates = sorted(list(self.blackout_dates))
            for date_str in sorted_dates:
                self.list_dates.addItem(date_str)

    def show_date_context_menu(self, position):
        item = self.list_dates.itemAt(position)
        if not item or item.text() == "æ— ":
            return
            
        menu = QMenu()
        delete_action = QAction("åˆ é™¤", self)
        delete_action.triggered.connect(lambda: self.delete_selected_date(item))
        menu.addAction(delete_action)
        menu.exec_(self.list_dates.mapToGlobal(position))
        
    def delete_selected_date(self, item):
        date_str = item.text()
        if date_str in self.blackout_dates:
            self.blackout_dates.remove(date_str)
            self.update_date_list()
            
    def init_advanced_tab(self):
        layout = QVBoxLayout(self.tab_advanced)
        
        # 1. Preferred Weekdays (æœŸæœ›å‘¨å‡ å€¼ç­)
        grp_weekdays = QGroupBox("1. æœŸæœ›å€¼ç­æ—¥ (æ¯å‘¨)")
        weekdays_layout = QHBoxLayout(grp_weekdays)
        self.weekday_checks = []
        days = ["å‘¨ä¸€", "å‘¨äºŒ", "å‘¨ä¸‰", "å‘¨å››", "å‘¨äº”", "å‘¨å…­", "å‘¨æ—¥"]
        preferred_days = self.preferences.get("preferred_weekdays", [])
        
        for i, day in enumerate(days):
            chk = QCheckBox(day)
            if i in preferred_days:
                chk.setChecked(True)
            self.weekday_checks.append(chk)
            weekdays_layout.addWidget(chk)
            
        layout.addWidget(grp_weekdays)

        # 2. Cycle Preference (å€¼ç­å‘¨æœŸ)
        grp_cycle = QGroupBox("2. åå¥½å€¼ç­å‘¨æœŸ")
        cycle_layout = QVBoxLayout(grp_cycle)
        self.combo_cycle = QComboBox()
        self.combo_cycle.addItems(["æ— ç‰¹å®šåå¥½", "æ¯å‘¨", "æ¯ä¸¤å‘¨ (éš”å‘¨)", "æ¯æœˆ"])
        
        # Load existing cycle
        current_cycle = self.preferences.get("preferred_cycle", "æ— ç‰¹å®šåå¥½")
        index = self.combo_cycle.findText(current_cycle)
        if index >= 0:
            self.combo_cycle.setCurrentIndex(index)
            
        cycle_layout.addWidget(self.combo_cycle)
        layout.addWidget(grp_cycle)

        # 3. Avoid Holidays (ä¸æœŸæœ›åœ¨å“ªä¸ªæ³•å®šèŠ‚å‡æ—¥å€¼ç­)
        grp_holiday = QGroupBox("3. ä¸æœŸæœ›å€¼ç­çš„èŠ‚å‡æ—¥")
        holiday_layout = QGridLayout(grp_holiday)
        self.holiday_checks = {}
        holidays = ["å…ƒæ—¦", "æ˜¥èŠ‚", "æ¸…æ˜èŠ‚", "åŠ³åŠ¨èŠ‚", "ç«¯åˆèŠ‚", "ä¸­ç§‹èŠ‚", "å›½åº†èŠ‚"]
        avoid_holidays = set(self.preferences.get("avoid_holidays", []))
        
        for i, h_name in enumerate(holidays):
            chk = QCheckBox(h_name)
            if h_name in avoid_holidays:
                chk.setChecked(True)
            self.holiday_checks[h_name] = chk
            holiday_layout.addWidget(chk, i // 4, i % 4)
            
        layout.addWidget(grp_holiday)
        
        # 4. Periodic Rotation (å®šæœŸè½®ç­)
        grp_rotation = QGroupBox("4. å®šæœŸè½®ç­ (ä¸ä»–äººè½®æµå€¼ç­)")
        rotation_layout = QFormLayout(grp_rotation)
        
        # Load existing rotation preference
        # Structure: {"partner": "CODE", "day_idx": 4, "parity": "odd"} 
        # parity: "odd" (1,3,5...) or "even" (2,4,6...)
        rotation_pref = self.preferences.get("periodic_rotation", {})
        
        # Partner Selector
        self.combo_rotation_partner = QComboBox()
        self.combo_rotation_partner.addItem("æ—  (ä¸å¯ç”¨)", None)
        
        current_partner_code = rotation_pref.get("partner")
        
        sorted_users = sorted(self.all_users, key=lambda u: u.code)
        for u in sorted_users:
            if u.id == self.user.id:
                continue
            self.combo_rotation_partner.addItem(f"{u.code} ({u.name or ''})", u.code)
            
        if current_partner_code:
            idx = self.combo_rotation_partner.findData(current_partner_code)
            if idx >= 0:
                self.combo_rotation_partner.setCurrentIndex(idx)
                
        rotation_layout.addRow("è½®ç­æ­æ¡£:", self.combo_rotation_partner)
        
        # Day Selector
        self.combo_rotation_day = QComboBox()
        
        # Parity Selector
        self.combo_rotation_parity = QComboBox()
        # odd = week 1, 3, 5...; even = week 2, 4, 6...
        self.combo_rotation_parity.addItem("å•å‘¨å€¼ç­ (ç¬¬1, 3, 5...å‘¨)", "odd")
        self.combo_rotation_parity.addItem("åŒå‘¨å€¼ç­ (ç¬¬2, 4, 6...å‘¨)", "even")
        
        current_parity = rotation_pref.get("parity", "odd")
        idx_parity = self.combo_rotation_parity.findData(current_parity)
        if idx_parity >= 0:
            self.combo_rotation_parity.setCurrentIndex(idx_parity)

        # Initialize days dynamically based on preferences
        self.update_rotation_days() 
        
        # Set initial value
        current_day = rotation_pref.get("day_idx", 4) # Default Friday
        idx = self.combo_rotation_day.findData(current_day)
        if idx >= 0:
            self.combo_rotation_day.setCurrentIndex(idx)
            
        rotation_layout.addRow("è½®ç­æ˜ŸæœŸ:", self.combo_rotation_day)
        rotation_layout.addRow("æˆ‘çš„ç­æ¬¡:", self.combo_rotation_parity)
        
        # Connect preferred weekdays change to rotation days update
        for chk in self.weekday_checks:
            chk.stateChanged.connect(self.update_rotation_days)

        if self.lock_rotation:
            p_idx = self.combo_rotation_partner.findData(self.lock_rotation["partner"])
            if p_idx >= 0:
                self.combo_rotation_partner.setCurrentIndex(p_idx)
            d_idx = self.combo_rotation_day.findData(self.lock_rotation["day_idx"])
            if d_idx >= 0:
                self.combo_rotation_day.setCurrentIndex(d_idx)
            r_idx = self.combo_rotation_parity.findData(self.lock_rotation["parity"])
            if r_idx >= 0:
                self.combo_rotation_parity.setCurrentIndex(r_idx)
            self.combo_rotation_partner.setEnabled(False)
            self.combo_rotation_day.setEnabled(False)
            self.combo_rotation_parity.setEnabled(False)
            lbl_locked = QLabel("è¯¥ç”¨æˆ·çš„è½®ç­ç”±æ­æ¡£è®¾ç½®ï¼Œå·²è‡ªåŠ¨é”å®šã€‚")
            lbl_locked.setStyleSheet("color: #FF9500;")
            rotation_layout.addRow(lbl_locked)

        # Explanation
        lbl_rot_hint = QLabel("è¯´æ˜ï¼šè®¾ç½®åï¼Œæ‚¨å°†ä¸æ­æ¡£åœ¨æŒ‡å®šæ˜ŸæœŸè½®æµå€¼ç­ã€‚\nè¯·ç¡®ä¿æ­æ¡£æœªè®¾ç½®å†²çªçš„è½®ç­è§„åˆ™ã€‚")
        lbl_rot_hint.setStyleSheet("color: gray; font-size: 11px;")
        rotation_layout.addRow(lbl_rot_hint)
        
        layout.addWidget(grp_rotation)
        
        layout.addStretch()

    def update_rotation_days(self):
        """Update the available rotation days based on preferred weekdays"""
        # Save current selection
        current_day_idx = self.combo_rotation_day.currentData()
        
        # Get preferred days indices
        preferred_indices = []
        for i, chk in enumerate(self.weekday_checks):
            if chk.isChecked():
                preferred_indices.append(i)
        
        self.combo_rotation_day.blockSignals(True)
        self.combo_rotation_day.clear()
        
        # If no preferred weekdays selected, disable rotation controls and show "None"
        if not preferred_indices:
            self.combo_rotation_day.addItem("æ— ", None)
            self.combo_rotation_day.setEnabled(False)
            self.combo_rotation_partner.setEnabled(False)
            self.combo_rotation_partner.setCurrentIndex(0) # Set to "None (Disabled)" or similar if exists, or just index 0
            self.combo_rotation_parity.setEnabled(False)
        else:
            # Enable controls
            self.combo_rotation_day.setEnabled(True)
            self.combo_rotation_partner.setEnabled(True)
            self.combo_rotation_parity.setEnabled(True)
            
            days = ["å‘¨ä¸€", "å‘¨äºŒ", "å‘¨ä¸‰", "å‘¨å››", "å‘¨äº”", "å‘¨å…­", "å‘¨æ—¥"]
            
            for i, day_name in enumerate(days):
                if i in preferred_indices:
                    self.combo_rotation_day.addItem(day_name, i)
            
            # Restore selection
            if current_day_idx is not None:
                idx = self.combo_rotation_day.findData(current_day_idx)
                if idx >= 0:
                    self.combo_rotation_day.setCurrentIndex(idx)
                else:
                    if self.combo_rotation_day.count() > 0:
                        self.combo_rotation_day.setCurrentIndex(0)
            elif self.combo_rotation_day.count() > 0:
                 self.combo_rotation_day.setCurrentIndex(0)
        
        self.combo_rotation_day.blockSignals(False)

    def get_preferences(self):
        prefs = self.preferences.copy()
        prefs["blackout_dates"] = sorted(list(self.blackout_dates))
        
        # New: Unavailable Weekdays
        unavailable_weekdays = []
        for i, chk in enumerate(self.unavailable_weekday_checks):
            if chk.isChecked():
                unavailable_weekdays.append(i)
        prefs["unavailable_weekdays"] = unavailable_weekdays
        
        # 1. Weekdays
        weekdays = []
        for i, chk in enumerate(self.weekday_checks):
            if chk.isChecked():
                weekdays.append(i)
        prefs["preferred_weekdays"] = weekdays
        
        # 2. Cycle
        prefs["preferred_cycle"] = self.combo_cycle.currentText()
        
        # 3. Holidays
        holidays = []
        for name, chk in self.holiday_checks.items():
            if chk.isChecked():
                holidays.append(name)
        prefs["avoid_holidays"] = holidays
        
        # 4. Periodic Rotation
        partner_code = self.combo_rotation_partner.currentData()
        if partner_code:
            prefs["periodic_rotation"] = {
                "partner": partner_code,
                "day_idx": self.combo_rotation_day.currentData(),
                "parity": self.combo_rotation_parity.currentData()
            }
        else:
            # If "None" is selected, remove the key if it exists
            if "periodic_rotation" in prefs:
                del prefs["periodic_rotation"]
        
        # Remove legacy pairing key if it exists, as UI is gone
        if "avoid_pairing" in prefs:
            del prefs["avoid_pairing"]

        # 5. Preferred Partners (New)
        # Collect from list_selected_partners
        if hasattr(self, 'list_selected_partners'):
            preferred_partners = []
            for i in range(self.list_selected_partners.count()):
                item = self.list_selected_partners.item(i)
                code = item.data(Qt.UserRole)
                if code:
                    preferred_partners.append(code)
            if preferred_partners:
                prefs["preferred_partners"] = preferred_partners
            elif "preferred_partners" in prefs:
                del prefs["preferred_partners"]
            
        return prefs

class UserEditDialog(QDialog):
    def __init__(self, user=None, parent=None):
        super().__init__(parent)
        self.user = user
        self.setWindowTitle("ç¼–è¾‘äººå‘˜" if user else "æ·»åŠ äººå‘˜")
        self.setFixedWidth(400)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        form_layout = QFormLayout()
        form_layout.setSpacing(15)
        
        # ID/Code
        self.edit_code = QLineEdit()
        if self.user:
            self.edit_code.setText(self.user.code)
            self.edit_code.setReadOnly(True) # Code is unique ID, usually not editable after creation
            self.edit_code.setPlaceholderText("ç³»ç»Ÿè‡ªåŠ¨ç”Ÿæˆæˆ–æ‰‹åŠ¨è¾“å…¥")
        else:
            self.edit_code.setPlaceholderText("å¿…å¡«ï¼Œå”¯ä¸€æ ‡è¯† (å¦‚ A, B, 001)")
            
        form_layout.addRow("å‘˜å·¥ID/ä»£ç :", self.edit_code)
        
        # Name
        self.edit_name = QLineEdit()
        self.edit_name.setPlaceholderText("å¿…å¡«ï¼Œæ˜¾ç¤ºåç§°")
        if self.user:
            self.edit_name.setText(self.user.name or self.user.code)
        form_layout.addRow("å§“å:", self.edit_name)
        
        # Position
        self.combo_position = QComboBox()
        self.combo_position.addItems(["å·¥é•¿", "å‰¯å·¥é•¿", "èŒå·¥", "è§ä¹ ç”Ÿ"])
        self.combo_position.setEditable(True) # Allow custom
        if self.user and self.user.position:
            self.combo_position.setCurrentText(self.user.position)
        form_layout.addRow("èŒåŠ¡:", self.combo_position)
        
        # Contact
        self.edit_contact = QLineEdit()
        if self.user and self.user.contact:
            self.edit_contact.setText(self.user.contact)
        form_layout.addRow("è”ç³»æ–¹å¼:", self.edit_contact)
        
        # Priority Level (formerly Employee Type)
        self.combo_employee_type = QComboBox()
        self.combo_employee_type.addItems(["ä¸€çº§", "äºŒçº§", "ä¸‰çº§"])
        current_type = "ä¸€çº§"
        if self.user and self.user.preferences:
            # Fallback to check permission_level if employee_type not set (migration support)
            raw_type = self.user.preferences.get("employee_type", "ä¸€çº§")
            # Map legacy values to new values
            mapping = {"ä¸€ç±»": "ä¸€çº§", "äºŒç±»": "äºŒçº§", "ä¸‰ç±»": "ä¸‰çº§"}
            current_type = mapping.get(raw_type, raw_type)
            
        self.combo_employee_type.setCurrentText(current_type)
        form_layout.addRow("ä¼˜å…ˆç­‰çº§:", self.combo_employee_type)
        
        # Color
        self.btn_color = QPushButton()
        self.current_color = self.user.color if self.user else "#3498DB"
        self.btn_color.setStyleSheet(f"background-color: {self.current_color}; border: none; border-radius: 4px;")
        self.btn_color.setFixedHeight(25)
        self.btn_color.clicked.connect(self.choose_color)
        form_layout.addRow("ä»£è¡¨é¢œè‰²:", self.btn_color)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        button_box.button(QDialogButtonBox.Save).setText("ä¿å­˜")
        button_box.button(QDialogButtonBox.Cancel).setText("å–æ¶ˆ")
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
    def choose_color(self):
        color = QColorDialog.getColor(initial=Qt.white, parent=self, title="é€‰æ‹©é¢œè‰²")
        if color.isValid():
            self.current_color = color.name()
            self.btn_color.setStyleSheet(f"background-color: {self.current_color}; border: none; border-radius: 4px;")
            
    def validate_and_accept(self):
        if not self.edit_code.text().strip():
            QMessageBox.warning(self, "éªŒè¯å¤±è´¥", "å‘˜å·¥ID/ä»£ç ä¸èƒ½ä¸ºç©º")
            return
        if not self.edit_name.text().strip():
            QMessageBox.warning(self, "éªŒè¯å¤±è´¥", "å§“åä¸èƒ½ä¸ºç©º")
            return
        self.accept()
        
    def get_data(self):
        return {
            "code": self.edit_code.text().strip(),
            "name": self.edit_name.text().strip(),
            "position": self.combo_position.currentText().strip(),
            "contact": self.edit_contact.text().strip(),
            "color": self.current_color,
            "employee_type": self.combo_employee_type.currentText()
        }
