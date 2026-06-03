import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    import xlwt
except ImportError:
    xlwt = None

class Exporter:
    def __init__(self, schedules, users):
        self.schedules = schedules
        self.users = users
        # Build dynamic map from user code to User object
        self.user_map = {u.code: u for u in users}

    def _get_daily_rows(self):
        """将排班记录按日期合并，返回每天的行数据"""
        daily_map = defaultdict(list)
        for sch in self.schedules:
            daily_map[sch.date].append(sch.user.code)
        
        # 排序日期
        sorted_dates = sorted(daily_map.keys())
        rows = []
        for i, date in enumerate(sorted_dates):
            user_codes = daily_map[date]
            # 确保有两个值班人员，不足补空
            u1_code = user_codes[0] if len(user_codes) > 0 else ""
            u2_code = user_codes[1] if len(user_codes) > 1 else ""
            
            u1_info = self._get_user_info(u1_code)
            u2_info = self._get_user_info(u2_code)
            
            weekday_map = {0: "一", 1: "二", 2: "三", 3: "四", 4: "五", 5: "六", 6: "日"}
            
            row_data = {
                "seq": i + 1,
                "date": date.strftime("%Y-%m-%d"),
                "weekday": weekday_map[date.weekday()],
                "time": "全天",
                "u1_name": u1_info["name"],
                "u2_name": u2_info["name"],
                "u1_phone": u1_info["phone"],
                "u2_phone": u2_info["phone"],
                "phone_display": u1_info["phone"] if u1_info["phone"] else "", 
                "remark": u2_info["phone"]
            }
            rows.append(row_data)
        return rows

    def _get_user_info(self, code):
        if not code:
            return {"name": "", "phone": ""}
        
        if code in self.user_map:
            user = self.user_map[code]
            # Use name if available, else code
            name = user.name if user.name else code
            # Use contact if available
            phone = user.contact if user.contact else ""
            return {"name": name, "phone": phone}
        else:
            return {"name": code, "phone": ""}

    def export_apple_style(self, filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "排班表"
        
        # 1. 标题
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Schedule Overview"
        title_cell.font = Font(name='Helvetica Neue', size=24, bold=True, color="333333")
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # 2. 表头
        headers = ["DATE", "DAY", "SHIFT", "STAFF 1", "STAFF 2", "CONTACT"]
        ws.append(headers)
        
        # 表头样式
        header_row = ws[2]
        for cell in header_row:
            cell.font = Font(name='Helvetica Neue', size=10, bold=True, color="888888")
            cell.fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(bottom=Side(style='thin', color='E0E0E0'))
        ws.row_dimensions[2].height = 25

        # 3. 数据内容
        rows = self._get_daily_rows()
        for i, row in enumerate(rows):
            data_row = [
                row['date'],
                "Wed" if row['weekday'] == "三" else "Thu" if row['weekday'] == "四" else "Fri" if row['weekday'] == "五" else "Sat" if row['weekday'] == "六" else "Sun" if row['weekday'] == "日" else "Mon" if row['weekday'] == "一" else "Tue", # 简单映射英文
                "All Day",
                row['u1_name'],
                row['u2_name'],
                row['phone_display']
            ]
            ws.append(data_row)
            
            # 数据行样式
            current_row = ws[i + 3]
            for cell in current_row:
                cell.font = Font(name='Helvetica Neue', size=12, color="333333")
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(bottom=Side(style='thin', color='EEEEEE'))
            
            ws.row_dimensions[i + 3].height = 30
            
        # 4. 列宽调整
        column_widths = [15, 10, 10, 15, 15, 20]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width
            
        # 去除网格线
        ws.views.sheetView[0].showGridLines = False
        
        wb.save(filepath)

    def export_to_excel(self, filepath, year=None, month=None):
        """Unified export method with year/month title support"""
        if filepath.lower().endswith('.xls'):
            return self._export_to_xls(filepath, year, month)
            
        wb = Workbook()
        ws = wb.active
        ws.title = "排班表"
        
        # 字体定义
        font_title = Font(name='宋体', size=20, bold=True)
        font_header = Font(name='宋体', size=11, bold=True)
        font_content = Font(name='宋体', size=11)
        
        border_style = Side(style='thin', color='000000')
        border_all = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        align_center = Alignment(horizontal='center', vertical='center')
        
        # 1. 大标题 (动态生成: "2025年6月排班表")
        title_text = "现场值班表"
        if year and month:
            title_text = f"{year}年{month}月排班表"
        elif year:
            title_text = f"{year}年排班表"
            
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = title_text
        title_cell.font = font_title
        title_cell.alignment = align_center
        title_cell.border = border_all
        ws.row_dimensions[1].height = 40
        
        # 2. 表头
        headers = ["序号", "日期", "星期", "时间", "值班1", "值班2", "值班电话", "备注"]
        ws.append(headers)
        
        header_row = ws[2]
        for cell in header_row:
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_all
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid") # 浅灰背景
            
        ws.row_dimensions[2].height = 25
        
        # 3. 数据填充
        rows = self._get_daily_rows()
        for i, row in enumerate(rows):
            data_row = [
                row['seq'],
                row['date'],
                row['weekday'],
                row['time'],
                row['u1_name'],
                row['u2_name'],
                row['phone_display'],
                row['remark']
            ]
            ws.append(data_row)
            
            current_row = ws[i + 3]
            for cell in current_row:
                cell.font = font_content
                cell.alignment = align_center
                cell.border = border_all
                
            ws.row_dimensions[i + 3].height = 22
            
        # 4. 列宽设置
        # 序号, 日期, 星期, 时间, 值班1, 值班2, 电话, 备注
        widths = [6, 15, 6, 8, 12, 12, 15, 15]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = w
            
        wb.save(filepath)

    def _export_to_xls(self, filepath, year=None, month=None):
        if xlwt is None:
            raise ImportError("xlwt 库未安装，无法导出 .xls 格式文件。")
            
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("排班表")
        
        # 字体定义
        font_title = xlwt.Font()
        font_title.name = '宋体'
        font_title.height = 20 * 20 # 20pt
        font_title.bold = True
        
        font_header = xlwt.Font()
        font_header.name = '宋体'
        font_header.height = 11 * 20
        font_header.bold = True
        
        font_content = xlwt.Font()
        font_content.name = '宋体'
        font_content.height = 11 * 20
        
        # 边框定义
        borders = xlwt.Borders()
        borders.left = xlwt.Borders.THIN
        borders.right = xlwt.Borders.THIN
        borders.top = xlwt.Borders.THIN
        borders.bottom = xlwt.Borders.THIN
        
        # 对齐方式
        align_center = xlwt.Alignment()
        align_center.horz = xlwt.Alignment.HORZ_CENTER
        align_center.vert = xlwt.Alignment.VERT_CENTER
        
        # 背景颜色
        pattern_header = xlwt.Pattern()
        pattern_header.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern_header.pattern_fore_colour = 22 # 浅灰背景 (22 对应 25% gray)
        
        # 样式对象
        style_title = xlwt.XFStyle()
        style_title.font = font_title
        style_title.alignment = align_center
        style_title.borders = borders
        
        style_header = xlwt.XFStyle()
        style_header.font = font_header
        style_header.alignment = align_center
        style_header.borders = borders
        style_header.pattern = pattern_header
        
        style_content = xlwt.XFStyle()
        style_content.font = font_content
        style_content.alignment = align_center
        style_content.borders = borders
        
        # 1. 大标题
        title_text = "现场值班表"
        if year and month:
            title_text = f"{year}年{month}月排班表"
        elif year:
            title_text = f"{year}年排班表"
            
        ws.write_merge(0, 0, 0, 7, title_text, style_title)
        ws.row(0).height_mismatch = True
        ws.row(0).height = int(40 * 20)
        
        # 2. 表头
        headers = ["序号", "日期", "星期", "时间", "值班1", "值班2", "值班电话", "备注"]
        for col_idx, h in enumerate(headers):
            ws.write(1, col_idx, h, style_header)
        ws.row(1).height_mismatch = True
        ws.row(1).height = int(25 * 20)
        
        # 3. 数据填充
        rows = self._get_daily_rows()
        for i, row in enumerate(rows):
            data_row = [
                row['seq'],
                row['date'],
                row['weekday'],
                row['time'],
                row['u1_name'],
                row['u2_name'],
                row['phone_display'],
                row['remark']
            ]
            row_idx = i + 2
            for col_idx, val in enumerate(data_row):
                ws.write(row_idx, col_idx, val, style_content)
                
            ws.row(row_idx).height_mismatch = True
            ws.row(row_idx).height = int(22 * 20)
            
        # 4. 列宽设置
        widths = [6, 15, 6, 8, 12, 12, 15, 15]
        for i, w in enumerate(widths):
            # xlwt width 约为 openpyxl 的 w * 256
            ws.col(i).width = int(w * 256)
            
        wb.save(filepath)

    def export_custom_style(self, filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "现场值班表"
        
        # 字体定义
        font_title = Font(name='宋体', size=20, bold=True)
        font_header = Font(name='宋体', size=11, bold=True)
        font_content = Font(name='宋体', size=11)
        
        border_style = Side(style='thin', color='000000')
        border_all = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        align_center = Alignment(horizontal='center', vertical='center')
        
        # 1. 大标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "现场值班表"
        title_cell.font = font_title
        title_cell.alignment = align_center
        title_cell.border = border_all
        ws.row_dimensions[1].height = 40
        
        # 2. 表头
        headers = ["序号", "日期", "星期", "时间", "值班1", "值班2", "值班电话", "备注"]
        ws.append(headers)
        
        header_row = ws[2]
        for cell in header_row:
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_all
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid") # 浅灰背景
            
        ws.row_dimensions[2].height = 25
        
        # 3. 数据填充
        rows = self._get_daily_rows()
        for i, row in enumerate(rows):
            data_row = [
                row['seq'],
                row['date'],
                row['weekday'],
                row['time'],
                row['u1_name'],
                row['u2_name'],
                row['phone_display'],
                row['remark']
            ]
            ws.append(data_row)
            
            current_row = ws[i + 3]
            for cell in current_row:
                cell.font = font_content
                cell.alignment = align_center
                cell.border = border_all
                
            ws.row_dimensions[i + 3].height = 22
            
        # 4. 列宽设置
        # 序号, 日期, 星期, 时间, 值班1, 值班2, 电话, 备注
        widths = [6, 15, 6, 8, 12, 12, 15, 15]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = w
            
        wb.save(filepath)
